import pandas as pd
from openpyxl import load_workbook

def export_sheets_to_csv(excel_file, output_directory):
    # Load the workbook
    workbook = load_workbook(filename=excel_file, data_only=True)
    
    for sheet_name in workbook.sheetnames:
        # Read each sheet into a DataFrame
        df = pd.DataFrame(workbook[sheet_name].values)
        
        # Set the output filename
        csv_filename = f"{output_directory}/{sheet_name}.csv"
        
        # Save to CSV
        df.to_csv(csv_filename, index=False, header=False)  # Update header/index as needed
        print(f"Exported: {csv_filename}")

# Example usage
excel_file = "data/all_tables.xlsx"  # Path to your Excel file
output_directory = "data"  # Directory to save CSV files
export_sheets_to_csv(excel_file, output_directory)
